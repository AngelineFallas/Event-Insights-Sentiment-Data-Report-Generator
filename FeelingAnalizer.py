import pandas as pd
import torch
import matplotlib.pyplot as plt
import seaborn as sns
from transformers import pipeline
import warnings
import time
from collections import Counter
from sentence_transformers import SentenceTransformer, util
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook

# Ignore warnings
warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=DeprecationWarning)

# Record the start time of execution
start_time = time.time()

# Load data
file_path = "volunteering_data(2).csv"
data = pd.read_csv(file_path, sep=";", encoding="utf-8", header=0)
data.columns = data.columns.str.strip()  # Remove leading/trailing spaces from column names

# Rename columns for consistency
data.rename(columns={
    "Select your role": "Role",
    "Enter your age": "Age",
    "What motivated you to participate?": "Motivation",
    "Briefly comment about your experience": "Comment",
    "What improvements should we apply to the program?": "Suggested Improvements",
    "Would you participate again?": "Would Participate Again",
    "Why would you participate again or not?": "Reason"
}, inplace=True)

# Fill missing values and convert Age to numeric
data.fillna("", inplace=True)
data["Age"] = pd.to_numeric(data["Age"], errors='coerce')
data.dropna(subset=["Age"], inplace=True)

# Load sentiment analysis model
sentiment_model = pipeline("sentiment-analysis", model="nlptown/bert-base-multilingual-uncased-sentiment")

# Function to classify sentiment labels
def analyze_sentiments(comments):
    results = sentiment_model(comments)
    sentiment_labels = {"1 star": "NEGATIVE", "2 stars": "NEGATIVE", "3 stars": "NEUTRAL", "4 stars": "POSITIVE", "5 stars": "POSITIVE"}
    return [sentiment_labels.get(res['label'], "NEUTRAL") for res in results]

# Apply sentiment analysis
valid_comments = data["Comment"].dropna().astype(str).tolist()
data["Sentiment"] = "NEUTRAL"
if valid_comments:
    data.loc[data["Comment"].notna(), "Sentiment"] = analyze_sentiments(valid_comments)

# Split data by role
groups = {role: df.copy() for role, df in data.groupby("Role")}

# Function to extract insights from a group
def extract_insights(df):
    return {
        "Motivations": Counter(df["Motivation"].dropna().astype(str)),
        "Positive Aspects": Counter(df.loc[df["Sentiment"] == "POSITIVE", "Reason"].dropna().astype(str)),
        "Negative Aspects": Counter(df.loc[df["Sentiment"] == "NEGATIVE", "Comment"].dropna().astype(str)),
        "Suggested Improvements": Counter(df["Suggested Improvements"].dropna().astype(str))
    }

# Create a new Excel workbook
wb = Workbook()
summary_ws = wb.active
summary_ws.title = "Analysis Summary"

# Add metadata to summary
now_str = time.strftime("%Y-%m-%d %H:%M:%S")
elapsed_time = round(time.time() - start_time, 2)
summary_ws.append(["Analysis Timestamp", now_str])
summary_ws.append(["Total Entries Analyzed", len(data)])
summary_ws.append(["Execution Time (seconds)", elapsed_time])
summary_ws.append([])

# Create role distribution pie chart in memory
role_counts = data["Role"].value_counts()
fig, ax = plt.subplots(figsize=(6, 6))
ax.pie(role_counts, labels=role_counts.index, autopct='%1.1f%%', colors=sns.color_palette("pastel"))
ax.set_title("Volunteer vs Donor Distribution")
img_data = BytesIO()
plt.savefig(img_data, format='png')
plt.close()
img_data.seek(0)
summary_ws.add_image(Image(img_data), "D2") #Excel cell where the upper corner of the image will paste

# Process each group (Volunteer, Donator)
for role, group_df in groups.items():
    ws = wb.create_sheet(title=f"{role} Report")
    insights = extract_insights(group_df)

    # Write each insight section with its own header
    for idx, (key, counter) in enumerate(insights.items()):
        ws.append([key])
        df = pd.DataFrame(counter.most_common(), columns=[key, "Count"])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        ws.append([])

    # Charts section
    chart_ws = wb.create_sheet(title=f"{role} Charts")

    # Sentiment bar chart
    sentiment_counts = group_df["Sentiment"].value_counts().reindex(["NEGATIVE", "NEUTRAL", "POSITIVE"], fill_value=0)
    fig, ax = plt.subplots(figsize=(6, 4))
    sns.barplot(x=sentiment_counts.index, y=sentiment_counts.values, palette="coolwarm", ax=ax)
    ax.set_title(f"Sentiment Distribution - {role}")
    ax.set_xlabel("Sentiment")
    ax.set_ylabel("Number of Comments")
    img = BytesIO()
    plt.savefig(img, format='png')
    plt.close()
    img.seek(0)
    chart_ws.add_image(Image(img), "B2") #Excel cell where the upper corner of the image will paste

    # Age group distribution
    group_df.loc[:, "Age Group"] = (group_df["Age"] // 10) * 10
    age_counts = group_df["Age Group"].value_counts().sort_index()
    fig, ax = plt.subplots(figsize=(6, 4))
    sns.barplot(x=age_counts.index.astype(str), y=age_counts.values, palette="pastel", ax=ax)
    ax.set_title(f"Age Distribution - {role}")
    ax.set_xlabel("Age Group (10-year intervals)")
    ax.set_ylabel("Count")
    img = BytesIO()
    plt.savefig(img, format='png')
    plt.close()
    img.seek(0)
    chart_ws.add_image(Image(img), "L2") #Excel cell where the upper corner of the image will paste

    # Participation pie chart
    participation_counts = group_df["Would Participate Again"].value_counts()
    fig, ax = plt.subplots(figsize=(6, 6))
    ax.pie(participation_counts, labels=participation_counts.index, autopct='%1.1f%%', colors=sns.color_palette("pastel"))
    ax.set_title(f"Would Participate Again - {role}")
    img = BytesIO()
    plt.savefig(img, format='png')
    plt.close()
    img.seek(0)
    chart_ws.add_image(Image(img), "G23") #Excel cell where the upper corner of the image will paste

# Save the final Excel file with timestamped name
file_time = time.strftime("%Y-%m-%d_%H-%M-%S")
filename = f"EventInsights_{file_time}.xlsx"
wb.save(filename)

print(f"Report saved to {filename}")
